import numpy as np
import openpyxl

def diagon(sq_matrix):
    pattern = np.zeros(sq_matrix.shape)
    for i in range(len(sq_matrix)-1):
        if not np.array_equal(pattern[i][i+1:], sq_matrix[i][i+1:]):
            return False
    return True

def get_matrix(filename):
    # wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    max_row = ws.max_row
    max_column = ws.max_column
    return np.array([ [(ws.cell(i, j)).value for j in range(1, max_column + 1)]
                    for i in range(1, max_row + 1)])

filename = 'input.xlsx'
print(diagon(get_matrix(filename)))