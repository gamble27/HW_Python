#!/usr/local/bin/python3.7
# -*- coding: utf-8 -*-

import cgi
from datetime import datetime as dt
from dateutil import parser
import os
import openpyxl

HTML_PAGE = """
<html>
<title>Решение задачи на рейсы</title>
<body>
<form method=POST action="">
<h1>Races</h1>
<font size="3" color="purple" face="Arial">
Аэропорт №1:
</font>
{0}
<br>
<font size="3" color="purple" face="Arial">
Аэропорт №2:
</font>
{1}
<br>
<font size="3" color="purple" face="Arial">
Введите дату:
</font>
<input type="date" id="start" name="chosen_date"
       value="{2}"
       min="{2}" max="{3}">
<br>
<input type=submit value="Отобразить список рейсов">
<br>

<table>
  <tr>
    <th>Date</th>
    <th>Airport #1</th>
    <th>Airport #2</th>
    <th>Depart</th>
    <th>Arrive</th>
    <th>Cost</th>
  </tr>
  {4}
</table>
</form>
</body>
</html>"""


class AirlinesOperations:

    def __init__(self, workbookio):

        self.wb = openpyxl.load_workbook(workbookio)

        self.dct = {}

        self._form_airport_dict()

    def _form_airport_dict(self):

        ws = self.wb["Аеропорти"]

        for i in range(1, ws.max_row + 1):
            self.dct[ws.cell(row=i, column=1).value] = [ws.cell(row=i, column=2).value + '\t(',
                                                        ws.cell(row=i, column=3).value + ')']

        del ws

    def server_ops(self):

        def application(environ, start_response):

            if environ.get('PATH_INFO', '').lstrip('/') == '':
                form = cgi.FieldStorage(fp=environ['wsgi.input'],
                                        environ=environ)

                result = ''

                try:
                    result = self.str_table(form.getvalue('AIR0').split()[0],
                                            form.getvalue('AIR1').split()[0],
                                            form.getvalue('chosen_date'))
                except AttributeError:
                    pass

                lst = [''.join(port) for port in self.dct.values()]

                body = HTML_PAGE.format(self.mkslct('AIR0', lst), self.mkslct('AIR1', lst),
                                        dt.now().strftime('%Y-%m-%d'), dt.max.strftime('%Y-%m-%d'),
                                        result)

                start_response('200 OK', [('Content-Type', 'text/html; charset=utf-8')])
            else:
                start_response('404 NOT FOUND', [('Content-Type', 'text/plain')])
                body = 'penguins'

            return [bytes(body, encoding='utf-8')]

        from wsgiref.simple_server import make_server

        print('=== my god.. ===')

        httpd = make_server('localhost', 4823, application)
        httpd.serve_forever()

    def str_table(self, from_port, to_port, dtime):

        to_port_id = from_port_id = ''

        for elem in self.dct.items():
            if from_port in elem[-1][0]:
                from_port_id = elem[0]
            elif to_port in elem[-1][0]:
                to_port_id = elem[0]

        ws0, ws1 = self.wb["Авіакомпанії"], self.wb["Рейси"]

        lst = []

        table_components = '''<tr>\n
                              <td>{}</td>\n
                              <td>{}</td>\n
                              <td>{}</td>\n
                              <td>{}</td>\n
                              <td>{}</td>\n
                              <td>{}</td>\n</tr>\n'''

        for i in range(1, ws1.max_row + 1):

            if (ws1.cell(row=i, column=1).value, ws1.cell(row=i, column=2).value) == (from_port_id, to_port_id):

                if str(parser.parse(dtime).weekday() + 1) in str(ws1.cell(row=i, column=4).value):

                    lst.append([''.join(self.dct[ws1.cell(row=i, column=1).value]),
                                ''.join(self.dct[ws1.cell(row=i, column=2).value]),
                                ws1.cell(row=i, column=5).value, ws1.cell(row=i, column=6).value,
                                ws1.cell(row=i, column=8).value])

        del ws0, ws1

        return ''.join(table_components.format(' '.join(dtime.split('-')[::-1]),
                                               *e) for e in lst)

    @staticmethod
    def mkslct(name, values):

        select_block = '<select name="{0}">\n{1}</select>\n'

        option_block = '<option value="{0}">{0}</option>\n'

        return select_block.format(name, ''.join(option_block.format(v) for v in values))

    @property
    def get_dict(self):
        return self.dct


if __name__ == '__main__':

    AirlinesOperations(os.path.join(os.getcwd(), 'airlines.xlsx')).server_ops()
